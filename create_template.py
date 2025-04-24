import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Medical Card Data"

# Define column headers
headers = ["Patient Name", "Age", "Sex", "Card Number", "Telephone", "Address", "Kebele", "Date", "Image Filename"]

# Add headers with formatting
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Set column widths
ws.column_dimensions["A"].width = 25  # Patient Name
ws.column_dimensions["B"].width = 15  # Age
ws.column_dimensions["C"].width = 15  # Sex
ws.column_dimensions["D"].width = 20  # Card Number
ws.column_dimensions["E"].width = 20  # Telephone
ws.column_dimensions["F"].width = 25  # Address
ws.column_dimensions["G"].width = 15  # Kebele
ws.column_dimensions["H"].width = 20  # Date
ws.column_dimensions["I"].width = 30  # Image Filename

# Add a few empty rows for data
for row in range(2, 12):
    ws.cell(row=row, column=1)
    ws.cell(row=row, column=2)
    ws.cell(row=row, column=3)
    ws.cell(row=row, column=4)
    ws.cell(row=row, column=5)
    ws.cell(row=row, column=6)
    ws.cell(row=row, column=7)
    ws.cell(row=row, column=8)
    ws.cell(row=row, column=9)

# Save the workbook
wb.save("template.xlsx")
print("Excel template created: template.xlsx") 